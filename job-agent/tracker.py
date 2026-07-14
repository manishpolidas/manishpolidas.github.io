#!/usr/bin/env python3
"""Job-application tracker: creates/updates job_applications.xlsx.

Usage: python3 tracker.py <apply|init> [json_file]
The JSON file holds a list of application-row dicts keyed by column name.
"""
import json
import os
import sys
from datetime import date, datetime, timedelta

from openpyxl import Workbook, load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "job_applications.xlsx")

APP_COLS = [
    "App ID", "Date Found", "Company", "Role Title", "Location / Remote",
    "Source", "JD URL", "Apply Method", "Match Score", "Resume Version",
    "Cover Letter", "Status", "Applied Date", "Blocked Reason",
    "Follow-up Date", "Recruiter / Contact", "Expected CTC", "Notes",
]
COMPANY_COLS = ["Company", "Sector", "City", "Careers URL", "ATS",
                "Date Added", "SAP-relevant"]
AUDIT_COLS = ["Timestamp", "App ID", "Company", "Role", "Resume Version",
              "Answers Used", "Action", "Detail"]


def get_wb():
    if os.path.exists(XLSX):
        return load_workbook(XLSX)
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(APP_COLS)
    wb.create_sheet("Summary")
    wb.create_sheet("Companies").append(COMPANY_COLS)
    wb.create_sheet("audit_log").append(AUDIT_COLS)
    return wb


def next_app_id(ws):
    year = date.today().year
    n = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0]).startswith(str(year)):
            n = max(n, int(str(row[0]).split("-")[1]))
    return f"{year}-{n + 1:03d}"


def rebuild_summary(wb):
    ws = wb["Applications"]
    counts = {}
    week_total = 0
    week_ago = (datetime.now() - timedelta(days=7)).date().isoformat()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        status = row[APP_COLS.index("Status")] or "?"
        counts[status] = counts.get(status, 0) + 1
        applied = row[APP_COLS.index("Applied Date")]
        if applied and str(applied)[:10] >= week_ago:
            week_total += 1
    s = wb["Summary"]
    s.delete_rows(1, s.max_row)
    s.append(["Status", "Count"])
    for k in sorted(counts):
        s.append([k, counts[k]])
    s.append([])
    s.append(["Applications submitted (last 7 days)", week_total])
    s.append(["Last updated", datetime.now().isoformat(timespec="seconds")])


def add_applications(rows):
    wb = get_wb()
    ws = wb["Applications"]
    ids = []
    for r in rows:
        app_id = r.get("App ID") or next_app_id(ws)
        r["App ID"] = app_id
        ws.append([r.get(c, "") for c in APP_COLS])
        ids.append(app_id)
    rebuild_summary(wb)
    wb.save(XLSX)
    return ids


def update_application(app_id, updates):
    wb = get_wb()
    ws = wb["Applications"]
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(app_id):
            for col, val in updates.items():
                row[APP_COLS.index(col)].value = val
            break
    rebuild_summary(wb)
    wb.save(XLSX)


def add_audit(entries):
    wb = get_wb()
    ws = wb["audit_log"]
    for e in entries:
        ws.append([e.get(c, "") for c in AUDIT_COLS])
    wb.save(XLSX)


def add_companies(rows):
    wb = get_wb()
    ws = wb["Companies"]
    existing = {str(r[0]).lower() for r in ws.iter_rows(min_row=2, max_col=1,
                                                        values_only=True) if r[0]}
    for r in rows:
        if r["Company"].lower() not in existing:
            ws.append([r.get(c, "") for c in COMPANY_COLS])
    wb.save(XLSX)


if __name__ == "__main__":
    cmd = sys.argv[1]
    payload = json.load(open(sys.argv[2])) if len(sys.argv) > 2 else None
    if cmd == "init":
        get_wb().save(XLSX)
        print("initialized", XLSX)
    elif cmd == "add":
        print(json.dumps(add_applications(payload)))
    elif cmd == "update":
        update_application(payload["app_id"], payload["updates"])
        print("updated", payload["app_id"])
    elif cmd == "audit":
        add_audit(payload)
        print("audited", len(payload))
    elif cmd == "companies":
        add_companies(payload)
        print("companies updated")
